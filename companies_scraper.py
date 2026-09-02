"""
Books / Companies Scraper
=========================
Scrapes product data from books.toscrape.com (public demo site)
and exports cleanly to JSON, CSV and Excel.

Author: Ahmad Raza
"""

import re
import json
import logging
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("books-scraper")

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)


def scrape_books(url: str = "https://books.toscrape.com") -> list[dict]:
    log.info("Starting scrape → %s", url)

    response = requests.get(url, timeout=20)
    response.raise_for_status()
    # Fix encoding so £ appears correctly
    response.encoding = response.apparent_encoding or "utf-8"

    soup = BeautifulSoup(response.text, "html.parser")
    books = soup.select("article.product_pod")

    data = []
    for book in books:
        name = book.h3.a.get("title", "").strip()
        name = re.sub(r"\s+", " ", name)

        price_el = book.select_one(".price_color")
        price = price_el.get_text(strip=True) if price_el else "N/A"

        rating_el = book.select_one("p.star-rating")
        rating = rating_el.get("class", ["", "N/A"])[1] if rating_el else "N/A"

        data.append({
            "name": name,
            "price": price,
            "rating": rating,
            "category": "Books",
            "source": "books.toscrape.com",
        })

    log.info("Scraped %d items", len(data))
    return data


def save_outputs(data: list[dict]):
    # JSON
    json_path = OUTPUT_DIR / "companies.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
    log.info("Saved %s", json_path)

    df = pd.DataFrame(data)

    # CSV
    csv_path = OUTPUT_DIR / "companies.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    log.info("Saved %s", csv_path)

    # Excel with nice formatting
    xlsx_path = OUTPUT_DIR / "companies.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Books"

        headers = list(df.columns)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 22
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
        log.info("Saved %s", xlsx_path)
    except Exception as e:
        # fallback
        df.to_excel(xlsx_path, index=False)
        log.warning("Basic Excel saved (formatting skipped): %s", e)

    print("\n--- Preview ---")
    print(df.head().to_string(index=False))
    print(f"\nTotal records: {len(df)}")


if __name__ == "__main__":
    try:
        records = scrape_books()
        save_outputs(records)
        log.info("Completed successfully")
    except Exception as e:
        log.error("Failed: %s", e)
        raise
